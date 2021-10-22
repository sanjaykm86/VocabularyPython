import openpyxl
import random
import requests
import schedule

path ='C:\\Users\\sanja\\OneDrive\\Vocabulary\\Vocabulary.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
def telegram_send(message):
    bot_token = '2075178780:AAGKfpTvO-shRKcxzbl9v2E6IxR1ZbXJ4Hg'
    bot_chat_id = '715032796'
    send_text = 'https://api.telegram.org/bot' + bot_token + '/sendMessage?chat_id=' + bot_chat_id + '&parse_mode=Markdown&text=' + message
    response = requests.get(send_text)


def sendVocabulary():
    count = sheet.max_row
    randomnumber = random.randrange(1,count)
    word = sheet.cell(row = randomnumber, column = 1).value
    meaning = sheet.cell(row=randomnumber,column=2).value
    example = sheet.cell(row= randomnumber, column=3).value
    # Print value of cell object
    # using the value attribute
    vocabularyMessage = f'''Word : {word}
    Meaning : {meaning}
    Example : {example}
    '''
    telegram_send(vocabularyMessage)


sendVocabulary()
#schedule.every(2).minutes.do(sendVocabulary)















 
  
